import os
import sys
import threading
import openpyxl
import requests
import logging
from bs4 import BeautifulSoup
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Font, Border, Side, PatternFill, NamedStyle
from openpyxl.chart import BarChart, Reference
import tkinter as tk
from tkinter import ttk


class SystemInfo:
    def __init__(self):
        self.bruger_navn = ""
        self.host_name = ""
        self.os_name = ""
        self.windows_version = ""
        self.sys_lang = ""
        self.bitlocker = ""


class DiskInfo:
    def __init__(self):
        self.fri_disk_plads_c = 0.0
        self.conversion_status = ""


# Create the progress bar window
root = tk.Tk()
root.title("Progress Window")
root.geometry("400x100")
progress = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
progress.pack()



logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    filename='FileConverter.log',  # Log messages are written to this file
                    filemode='w')  # 'w' for overwrite mode, 'a' for append mode

def count_specific_files_in_directory(directory_path, extension):
    """Count the number of files with a specific extension in the given directory."""
    return sum(1 for entry in os.listdir(directory_path)
               if os.path.isfile(os.path.join(directory_path, entry))
               and entry.endswith(extension))

# Update the progress bar
def update_progress(value):
    def _update():
        progress['value'] = value
    root.after(0, _update)

def extract_data(file_path):
    data = SystemInfo()
    volume_data = DiskInfo()
    processing_volume = False

    try:
        with open(file_path, 'r', encoding='ansi') as file:
            lines = file.readlines()
            data.bruger_navn = lines[0].strip()  # Set username

            free_disc_space = lines[-1].strip()
            if free_disc_space.startswith('2 Verzeichnis(se)'):
                data.fri_disk_plads_c = float(free_disc_space[19:24])
            else:
                data.fri_disk_plads_c = float(
                    free_disc_space[10:15])  # Set free disk space

            for line in lines:
                if processing_volume:
                    if line.startswith('    Conversion Status'):
                        volume_data.conversion_status = line.strip().split(':')[
                            1].strip()
                        processing_volume = False  # Done finding conversion status
                else:
                    if line.startswith('Volume C:'):
                        processing_volume = True
                        volume_data = DiskInfo()  # Reset volume_data
                    elif line.startswith('OS Version:') or line.startswith('Version du systŠme'):
                        win_version = line.strip().split(':')[1].strip().split()
                        data.windows_version = ' '.join(
                            [win_version[0], win_version[2], win_version[3]])
                    elif line.startswith('Betriebssystemversion'):
                        win_version = line.strip().split(':')[1].strip().split()
                        data.windows_version = ' '.join(
                            [win_version[0], win_version[3], win_version[4]])
                    elif line.startswith('Host Name:') or line.startswith('Hostname:') or line.startswith('Nom de l'):
                        data.host_name = line.strip().split(
                            ':')[1].strip()  # Set Host Name
                    elif line.startswith('OS Name:') or line.startswith('Betriebssystemname:') or line.startswith(
                            'Nom du systŠme'):
                        data.os_name = line.strip().split(
                            ':')[1].strip()  # Set OS Name
                    elif line.startswith('System Locale') or line.startswith('Systemgebietsschema') or line.startswith(
                            'Option r‚gionale du systŠme'):
                        data.sys_lang = line.strip().split(':')[1].strip()

            if volume_data.conversion_status:
                data.bitlocker = volume_data.conversion_status  # Save BitLocker information

            print("Windows 11:", win11_version)
            print("Windows 10:", win10_version)
            print("Data:", data.windows_version)
            test = data.windows_version.strip().split(" ")[2].strip()
            print("Test: ", test)
            if test == win11_version or data.windows_version == win10_version:
                data.is_up_to_date = True

        return data
    except Exception as e_extractData:
        logging.error(f"Error processing file {file_path}: {e_extractData}")


def convert_french_format_to_number(french_format):
    # Remove "octets libres" from the string and replace "ÿ" with nothing
    cleaned_format = french_format.replace("ÿ", ",")
    return cleaned_format


def get_latest_windows_versions(url):
    # Fetch content
    try:
        response = requests.get(url)
        html_content = response.content

        soup = BeautifulSoup(html_content, "html.parser")

        # Find elements with the latest version and build
        target_element = ""

        if url.__contains__("windows11"):
            target_element = soup.select(
                "html.hasSidebar.hasPageActions.hasBreadcrumb.conceptual.has-default-focus.theme-light"
                " body div.mainContainer"
                ".uhf-container.has-default-focus div.columns.has-large-gaps.is-gapless-mobile "
                "section.primary-holder.column.is-two-thirds-tablet.is-three-quarters-desktop div.columns."
                "is-gapless-mobile.has-large-gaps"
                " div#main-column.column.is-full.is-8-desktop main#main div.content div#winrelinfo_container strong")

        elif url == "https://learn.microsoft.com/en-us/windows/release-health/release-information":
            target_element = soup.select(
                "html.hasSidebar.hasPageActions.hasBreadcrumb.conceptual.has-default-focus.theme-light body "
                "div.mainContainer.uhf-container.has-default-focus div.columns.has-large-gaps.is-gapless-mobile "
                "section.primary-holder.column.is-two-thirds-tablet.is-three-quarters-desktop"
                " div.columns.is-gapless-mobile.has-large-gaps "
                "div#main-column.column.is-full.is-8-desktop main#main div.content div#winrelinfo_container strong")

        if target_element:
            # Get data from target_element
            content = target_element[0].get_text().strip()
            return content
        else:
            logging.error("The Windows Version could not be found / Target not found")
    except requests.exceptions.RequestException as e:
        logging.error(f"Network error occurred: {e}")
        return None


win11 = "https://learn.microsoft.com/en-us/windows/release-health/windows11-release-information"
win10 = "https://learn.microsoft.com/en-us/windows/release-health/release-information"

try:
    win11_version = get_latest_windows_versions(win11).split(" ")[4].split(")")[0]
    win10_version = get_latest_windows_versions(win10).split(" ")[4].split(")")[0]
except Exception as e:
    logging.error("There has been an error retrieving the the Windows website")
    logging.error("Make sure the computer is connected to the internet.")
    logging.exception("Error: %s", e)


def get_executable_directory():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))


def main(update_progress_callback):
    folder_path = get_executable_directory()
    logging.info("Main Function Started")

    # Count the total number of .txt files in the directory
    total_files = count_specific_files_in_directory(folder_path, '.txt')
    if total_files == 0:
        logging.error("No .txt files found in the directory.")
        return

    # Start excel book
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Windows Update Status'
    logging.info("Excel document created")

    # Border style
    medium_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    # Bold text
    bold_font = Font(bold=True)

    # Write data names
    worksheet.append(['Bruger Navn',
                      'Host Navn',
                      'OS Navn',
                      'Windows Version',
                      'Fri disk plads på C:',
                      'BitLocker Status'])

    number_format_style = NamedStyle(
        name='number_format_style',
        number_format='##0,0')
    target_column = "E"
    worksheet[f'{target_column}1'].style = number_format_style

    try:
        file_count = 0
        for filename in os.listdir(folder_path):
            if filename.endswith('.txt'):
                logging.info("Getting info from files in path: %s", filename)
                file_path = os.path.join(folder_path, filename)
                if os.path.isfile(file_path):
                    data = extract_data(file_path)
                    file_count += 1
                    update_progress(100 * file_count / total_files)

                    # Write data to Excel file
                    row = [data.bruger_navn,
                           data.host_name,
                           data.os_name,
                           data.windows_version,
                           data.fri_disk_plads_c,
                           data.bitlocker]
                    worksheet.append(row)

                    # Data for disk space is in the 5th column and the usernames are in the 1st column
                    # Reference for chart data - starting from the second row to skip the header
                    chart_data = Reference(worksheet, min_col=5, min_row=2, max_row=worksheet.max_row)
                    # Reference for categories - usernames
                    categories = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)

                    # Create a bar chart
                    chart = BarChart()
                    chart.add_data(chart_data, titles_from_data=False)
                    chart.set_categories(categories)
                    chart.title = "Free Disk Space on C:"
                    chart.y_axis.title = 'Disk Space (GB)'
                    chart.x_axis.title = 'Bruger'
                else:
                    logging.error(f"File not found: {file_path}")
                root.after(0, update_progress_callback, 100 * file_count / total_files)

            else:
                logging.info("Filename does not end with .txt: %s", filename)

        # Place the chart on the worksheet
        worksheet.add_chart(chart, "G10")  # Update cell location as needed

        header_row = worksheet[1]
        for cell in header_row:
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thick'))

        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Define different colors for cells
        green_fill = PatternFill(
            start_color="00FF00",
            end_color="00FF00",
            fill_type="solid")
        red_fill = PatternFill(
            start_color="FF0000",
            end_color="FF0000",
            fill_type="solid")
        yellow_fill = PatternFill(
            start_color="E4CD00",
            end_color="E4CD00",
            fill_type="solid")

        for row in worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column):
            os_name = row[2].value
            win_version_row = row[3].value
            win_version = win_version_row[-5:]

            if os_name == "Microsoft Windows 11 Pro":
                if win_version == win11_version:
                    row[3].fill = green_fill
                else:
                    row[3].fill = red_fill
                    for i in range(4):
                        row[i].border = medium_border
            elif os_name == "Microsoft Windows 10 Pro":
                if win_version == win10_version:
                    row[3].fill = green_fill
                else:
                    row[3].fill = red_fill
                    for i in range(4):
                        row[i].border = medium_border
            else:
                row[3].fill = yellow_fill

        disk_space_rule = DataBarRule(start_type="num",
                                      start_value=1,
                                      end_type="num",
                                      end_value="300",
                                      color="0000FF00")  # Green
        worksheet.conditional_formatting.add("E2:E1000", disk_space_rule)

        for row_num in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_num, column=5)
            if cell.value is not None and cell.value < 50:
                cell.font = bold_font

        for row_num in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_num, column=6)
            if cell.value == 'Fully Encrypted':
                cell.fill = PatternFill(
                    start_color="00FF00",
                    end_color="00FF00",
                    fill_type="solid")
            else:
                cell.fill = PatternFill(
                    start_color="FF0000",
                    end_color="FF0000",
                    fill_type="solid")

        worksheet.freeze_panes = "C2"

        # Save Excel file
        excel_filename = '0_samlet_data.xlsx'
        excel_path = os.path.join(folder_path, excel_filename)
        workbook.save(excel_path)

        logging.info(f"Data retrieved and saved to: {excel_filename}")

    except Exception as exception:
        logging.error("An error has occured while running the program.")
        logging.exception("An error occurred: %s", exception)
        input("Press Enter to close the window...")

def start_main_thread():
    threading.Thread(target=main, args=(update_progress,)).start()

if __name__ == "__main__":
    start_main_thread()

    root.mainloop()
